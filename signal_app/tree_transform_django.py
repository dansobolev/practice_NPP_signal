from .trees_algorithms import to_descendants_list
import pandas as pd

from openpyxl.styles import PatternFill
from openpyxl import load_workbook


# Отображение: дец. номер -> id (id - номер строки в БД)
def id_dict(assembly, basep):
    str_count = 0  # Счетчик количества срок
    d = dict()  # Отображение: дец. номер -> id
    for i in assembly:
        d[i.decimal_number] = str_count
        str_count +=1
    for i in basep:
        d[i.decimal_number] = str_count
        str_count +=1
    return d


# Пребразование БД в список непосредственных предков
def ancestors_list(assembly, basep):
    d = id_dict(assembly, basep)
    ancs_lst = [-1]  # TODO: С учетом того, что крень дерева первый в табилце assembly
    for j in [assembly[1:], basep]:
        for i in j:
            ancs_lst.append(d[i.entry_number])
    return ancs_lst


# Преобразование БД во вложенный словарь
def bd_to_dict(assembly, basep, k=0, descs_lst=None):
    if descs_lst is None:
        ancs_lst = ancestors_list(assembly, basep)
        descs_lst = to_descendants_list(ancs_lst)
    # Формирование словаря для текущей детали
    # names = ['id', 'name', 'amount', 'vhod', 'type']
    names = ['id', 'name', 'vhod', 'type']
    n = len(assembly)
    if k < n:
        values = [assembly[k].decimal_number, assembly[k].name, assembly[k].entry_number, 0]
    else:
        values = [basep[k-n].decimal_number, basep[k-n].name, basep[k-n].entry_number, basep[k-n].product_type]
    d = {n: v for n, v in zip(names, values)}
    d['sub_assembly'] = []
    d['sub_details'] = []
    # Формирование потомков для текущей детали
    for i in descs_lst[k]:
        if i < n:
            d['sub_assembly'].append(bd_to_dict(assembly, basep, i, descs_lst))
        else:
            d['sub_details'].append(bd_to_dict(assembly, basep, i, descs_lst))
    # Возвращает словарь изделия, список предков, список потомков
    return d


def find_path(k, ancs_lst, descs_lst):
    if ancs_lst[k] != -1:
        return [k] + find_path(ancs_lst[k], ancs_lst, descs_lst)
    else:
        return [k]


def add_edge(id, name, vhod, type, product_dict, ancestors_list, descendants_list, id_dict):
    '''
    param: id - Децимальный номер добалвяеомго изделия
    param: name - Наименование добавляемого изделия
    param: vhod - Дециальный номер предка (первичная входимость)
    param: type - Тип добавляемого объекта (assembly, detail)
    param: product_dict - Словарь изделия
    param: ancestors_list - Список непосредственных предков дерева
    param: descendants_list - Список непосредственных потомков дерева
    param: id_dict - Отображение (дец. номер -> номер вершины)
    '''
    k = len(ancestors_list)  # Номер добавляемой вершины
    # Обновление информации о дереве
    p_id_dict = id_dict.copy()  # Обновление словаря (дец. номер -> номер вершины)
    p_id_dict[id] = k
    ancs_lst = ancestors_list + [p_id_dict[vhod]]  # Обновление списка непосредственных предков
    descs_lst = descendants_list + [[]]  # Обновление списка непосредственных потомков
    descs_lst[p_id_dict[vhod]].append(k)
    # Добавление вершины в словарь изделия
    prod_dict = product_dict.copy()
    t = prod_dict
    path = list(reversed(find_path(k, ancs_lst, descs_lst)))[1:-1]
    for i in path:
        j = list(map(lambda x: p_id_dict[x['id']], t['sub_assembly'])).index(i)
        t = t['sub_assembly'][j]
    current_dict = {'id': id, 'name': name, 'vhod': vhod, 'type': type, 'sub_assembly': [], 'sub_details': []}
    if type:  # Если type != 0, то добавляется деталь
        t['sub_details'].append(current_dict)
    else:
        t['sub_assembly'].append(current_dict)
    return prod_dict


def dict_to_table(product_dict):
    table = [[product_dict['id'], product_dict['name'], product_dict['vhod'], product_dict['type']]]
    if product_dict['sub_details']:  # Если список деталей не пуст
        for i in product_dict['sub_details']: # TODO: Полагаем, что внутри деталей нет других деталей
            table.append([i['id'], i['name'], i['vhod'], i['type']])
    if product_dict['sub_assembly']:  # Если список подсборок не пуст
        for i in product_dict['sub_assembly']:
            table.extend(dict_to_table(i))
    return table


# Список потомков по словарю
# TODO: Функция не работает
def dict_to_descs_lst(product_dicts, edge_c=1, id_dict=None):
    if id_dict is None:
        id_dict = dict()
    if product_dicts:
        descs_lst = []  # Список потомков (Номера) для всех вершин на данном уровне
        for i in product_dicts:  # Рассматриваем конкретную вершину
            descs_lst_d = []  # Список потомков (словари) для конкретной вершины
            descs_lst_n = []  # Список потомков (номера) для конкретной вершины
            # Сборки (потомки)
            assembly_descs_d = []
            assembly_descs_n = []
            for j in i['sub_assembly']:  # Рассматриваем конкретного потомка вершины
                assembly_descs_d.append(j)  # Добалвение словаря потомка
                assembly_descs_n.append(edge_c)  # Добалвение номера потомка
                id_dict[i['id']] = edge_c  # Добалвение отображения (дец. ном. потомка -> ном. вершины)
                edge_c += 1
            descs_lst_d.extend(assembly_descs_d)
            descs_lst_n.extend(assembly_descs_n)
            # Детали (потомки)
            detail_descs_d = []
            detail_descs_n = []
            for j in i['sub_details']:  # Рассматриваем конкретного потомка вершины
                detail_descs_d.append(j)  # Добалвение словаря потомка
                detail_descs_n.append(edge_c)  # Добалвение номера потомка
                id_dict[i['id']] = edge_c  # Добалвение отображения (дец. ном. потомка -> ном. вершины)
                edge_c += 1
            descs_lst_d.extend(detail_descs_d)
            descs_lst_n.extend(detail_descs_n)
            descs_lst.append(descs_lst_n)
            return descs_lst
    else:
        return []
    return 'error'


# TODO: Реализуй удаление через формирование словаря и всего остального по готовой таблице
def delete_edge(id, type, product_dict, ancs_lst, descs_lst, id_dict, assembly_length):
    '''
    param: id - Децимальный номер добалвяеомго изделия
    param: type - Тип удаляемой вершины
    param: product_dict - Словарь изделия
    param: ancs_lst - Список непосредственных предков дерева
    param: descs_lst - Список непосредственных потомков дерева
    param: id_dict - Отображение (дец. номер -> номер вершины)
    '''
    # Удаление вершины из словаря
    k = id_dict[id]  # Номер удаляемой вершины
    prod_dict = product_dict.copy()
    t = prod_dict
    path = list(reversed(find_path(k, ancs_lst, descs_lst)))[1:]
    for i in path[:-1]:
        j = list(map(lambda x: id_dict[x['id']], t['sub_assembly'])).index(i)
        t = t['sub_assembly'][j]
    if type:  # Если type != 0, то удаляется деталь
        j = list(map(lambda x: id_dict[x['id']], t['sub_details'])).index(path[-1])
        deleted = t['sub_details'][j].copy()
        # t['sub_details'].pop(j)
        return [], dict_to_table(deleted)
    else:
        j = list(map(lambda x: id_dict[x['id']], t['sub_assembly'])).index(path[-1])
        deleted = t['sub_assembly'][j].copy()
        # t['sub_assembly'].pop(j)
        tab = dict_to_table(deleted)
        assembly_lst, detail_lst = [], []
        for i in tab:
            if i[-1]:
                detail_lst.append([i[1], i[2]])
            else:
                assembly_lst.append(i[0])
        return assembly_lst, detail_lst  # assembly_lst - список децимальников (Assembly), list of ids (BaseProduct)


def change_edge(id, type, name, product_dict, ancs_lst, descs_lst, id_dict, id_c=None, name_c=None, type_c=None):
    '''
    param: id - Децимальный номер изменяемой вершины
    param: type - Тип удаляемой вершины
    param: product_dict - Словарь изделия
    param: ancs_lst - Список непосредственных предков дерева
    param: descs_lst - Список непосредственных потомков дерева
    param: id_dict - Отображение (дец. номер -> номер вершины)
    param: id_c - Новый дец. номер вершины
    param: name_c - Новое имя вершины
    param: type_c - Новый тип вершины (если изменяется деталь)
    '''
    if id_c is None:
        id_c = id
    if name_c is None:
        name_c = name
    if type_c is None:
        type_c = type
    # Изменение вершины
    k = id_dict[id]  # Номер изменяемой вершины
    prod_dict = product_dict.copy()
    t = prod_dict
    path = list(reversed(find_path(k, ancs_lst, descs_lst)))[1:]
    for i in path[:-1]:
        j = list(map(lambda x: id_dict[x['id']], t['sub_assembly'])).index(i)
        t = t['sub_assembly'][j]
    if type:  # Если type != 0, то изменяется деталь
        j = list(map(lambda x: id_dict[x['id']], t['sub_details'])).index(path[-1])
        changed = t['sub_details'][j].copy()
        return [], [{'name': changed['name'], 'entry_number': changed['vhod'],
                'fields_to_edit': {'name': name_c, 'decimal_number': id_c, 'product_type': type_c}}]
    else:
        j = list(map(lambda x: id_dict[x['id']], t['sub_assembly'])).index(path[-1])
        changed = t['sub_assembly'][j].copy()
        tab = [[changed['id'], changed['name'], changed['vhod'], changed['type']],
               id_c, name_c, changed['vhod'], changed['type']]
        # Изменение входимости потомков
        # потомки в подсборках
        assembl = []  # Что меняется
        for i in changed['sub_assembly']:
            assembl.append({'decimal_number': i['id'], 'fields_to_edit': {'entry_number': id_c}})
        # Потомки в деталях
        detail = []  # Что меняется
        for i in changed['sub_details']:
            detail.append({'name': i['name'], 'entry_number': i['vhod'], 'fields_to_edit': {'entry_number': id_c}})
    return assembl, detail


def save_dict_to_excel(product_dict):
    # Для работы функции необходим модуль openpyxl
    tab = dict_to_table(product_dict)
    cols = ['Децимальный номер', 'Наименование', 'Первичная входимость', 'Принадлежность']
    df = pd.DataFrame(tab, columns=cols)
    type_dict = {
        0: 'Сборка',
        1: 'Деталь',
        2: 'Стандартное изделие',
        3: 'Прочие изделия',
        'assembly': 'Сборка'}
    color_dict = {
        'Сборка': '93C47D',
        'Деталь': 'E2EFDA',
        'Стандартное изделие': 'FCE4D6',
        'Прочие изделия': 'D9E1F2'
    }
    df['Принадлежность'] = df['Принадлежность'].map(type_dict)
    df.to_excel('output.xlsx', index=False)
    wb = load_workbook('output.xlsx')
    ws = wb['Sheet1']
    # Форматирование
    for cell in list(ws)[1]:
        cell.fill = PatternFill(start_color='38761D',
                                end_color='38761D',
                                fill_type='solid')
    for row in list(ws)[2:]:
        color = color_dict[row[-1].value]
        for cell in row:
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type='solid')
    wb.save('Изделие1.xlsx')
    return 'Saved'
