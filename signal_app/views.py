import ast
from datetime import datetime
import json

from django.shortcuts import render, HttpResponse
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from users.models import UserProfile
from .models import Assembly, BaseProduct
from .enums import TypeEnum
from .exceptions import ObjectDoesNotExist
from .tree_transform_django import bd_to_dict, dict_to_table, ancestors_list, id_dict, delete_edge, change_edge
from users.permissions import project_permissions_required, ProjectPermissions
from .trees_algorithms import to_descendants_list


def get_all_items_from_bd():
    return Assembly.objects.all(), BaseProduct.objects.all()


def bd_to_dict_helper(check_empty_bd: bool = False):
    assembly, base_products = get_all_items_from_bd()
    if check_empty_bd:
        if assembly.count() == 0 and base_products.count() == 0:
            return JsonResponse({"": ""})

    return bd_to_dict(assembly, base_products)


def index(request):
    return render(request, 'base.html')


@login_required
@project_permissions_required(
    permissions=[ProjectPermissions.PERM_PROJECT_UPDATE, ProjectPermissions.PERM_PROJECT_READ]
)
def show_tree(request):
    response = bd_to_dict_helper(check_empty_bd=True)

    return JsonResponse(response)


@login_required
@project_permissions_required(permissions=[ProjectPermissions.PERM_PROJECT_UPDATE])
def save_data(request):
    body = json.loads(request.body)
    type_ = TypeEnum(int(body['type']))
    body.update({'type': type_.name.title()})

    if type_.name != TypeEnum.ASSEMBLY.name:
        try:
            Assembly.objects.filter(decimal_number=body['vhod']).get()
        except Assembly.DoesNotExist:
            return JsonResponse({"message": "Данная сборка отсутствует в базе. Перепроверьте данные."})

        new_db_object = BaseProduct.objects.create(
            decimal_number=body['number'],
            name=body['name'],
            entry_number=body['vhod'],
            count_number=3,
            product_type=type_.value
        )
        new_db_object.save()

        return HttpResponse(request, status=204)

    new_db_object = Assembly.objects.create(
        decimal_number=body['number'],
        name=body['name'],
        entry_number=body['vhod'],
    )
    new_db_object.save()

    return JsonResponse({"status_code": 200, "message": "Item has been successfully added to DB"})


@login_required
@project_permissions_required(permissions=[ProjectPermissions.PERM_PROJECT_UPDATE])
def delete_entity(request):
    data = json.loads(request.body)
    item_type = TypeEnum(int(data['type']))
    if 'decimal_number' in data:
        assembly, base_products = get_all_items_from_bd()
        product_dict = bd_to_dict_helper(check_empty_bd=False)
        ancs_list = ancestors_list(assembly, base_products)
        descs_list = to_descendants_list(ancs_list)
        id_d = id_dict(assembly, base_products)

        assemblies_to_delete, base_products_to_delete = delete_edge(
            data['decimal_number'], item_type.value, product_dict, ancs_list, descs_list, id_d, len(assembly)
        )
        try:
            if not data['decimal_number']:
                product_name = data['name']
                product_entry_number = data['entry_number']
                product = BaseProduct.objects.filter(
                    name=product_name, entry_number=product_entry_number,
                ).get()
                product.delete()
                return JsonResponse({"status_code": 200, "message": "Item has been successfully deleted."})

            # удаление сборок по децимальникам
            if assemblies_to_delete:
                for assembly_decimal in assemblies_to_delete:
                    assembly = Assembly.objects.filter(decimal_number=assembly_decimal).get()
                    assembly.delete()
                # удаление подсборок по id в БД
                for item in base_products_to_delete:
                    name = item[0]
                    entry_number = item[1]
                    product = BaseProduct.objects.filter(name=name, entry_number=entry_number).get()
                    product.delete()
            else:
                for item in base_products_to_delete:
                    name = item[1]
                    entry_number = item[2]
                    product = BaseProduct.objects.filter(name=name, entry_number=entry_number).get()
                    product.delete()
        except Exception as e:
            return JsonResponse({"error": "Cannot find item with provided decimal number."})

    return JsonResponse({"status_code": 200, "message": "Item has been successfully deleted."})


@login_required
@project_permissions_required(permissions=[ProjectPermissions.PERM_PROJECT_UPDATE])
def edit_entity(request):
    data = json.loads(request.body)
    item_type = TypeEnum(int(data['type']))
    print("data from front: ", data)
    if 'decimal_number' and 'fields_to_edit' in data:
        assembly, base_products = get_all_items_from_bd()
        product_dict = bd_to_dict_helper(check_empty_bd=False)
        ancs_list = ancestors_list(assembly, base_products)
        descs_list = to_descendants_list(ancs_list)
        id_d = id_dict(assembly, base_products)
        #old_name = data['old_name']
        old_name = 'Крыша'

        id_c = data['fields_to_edit']['decimal_number'] if 'decimal_number' in data['fields_to_edit'] else None
        name_c = data['fields_to_edit']['name'] if 'name' in data['fields_to_edit'] else None
        type_c = data['fields_to_edit']['type'] if 'type' in data['fields_to_edit'] else None

        assemblies_edit, base_products_edit = change_edge(
            id=data['decimal_number'],
            type=item_type.value,
            name=old_name,
            product_dict=product_dict,
            ancs_lst=ancs_list,
            descs_lst=descs_list,
            id_dict=id_d,
            id_c=id_c, name_c=name_c, type_c=type_c,
        )

        # TODO: изменять имя самому ( для сборки )
        # TODO: изменять имя у всех деталей (если другие поля не меняются)

        # TODO: сначала изменить первичную входимость у деталей, потом у сборки изменить децимальник

        # если у элемента изменяется только поле name - меняем его сразу
        if 'name' in data['fields_to_edit']:
            print("HERE")
            if item_type == TypeEnum.ASSEMBLY:
                item_to_change = Assembly.objects.filter(decimal_number=data['decimal_number']).get()
                item_to_change.name = data['fields_to_edit']['name']
                item_to_change.save()
            else:
                item_to_change = BaseProduct.objects.filter(
                    name=data['old_name'], entry_number=data['entry_number']
                ).get()
                item_to_change.name = data['fields_to_edit']['name']
                item_to_change.save()
            return JsonResponse({"status_code": 200, "message": "Item has been successfully edited."})

        for assembly_to_edit in assemblies_edit:
            old_decimal_number = data['decimal_number']
            print('QQQQQQ')

            # случай когда у предка меняется децимальник и его надо изменить и у первичной входимости потомков
            assembly_decimal_number_to_edit = assembly_to_edit.get('decimal_number', None)
            if assembly_decimal_number_to_edit is not None:
                try:
                    assembly_to_change = Assembly.objects.filter(
                        decimal_number=old_decimal_number
                    ).get()
                except Exception as e:
                    return JsonResponse({"error": "Cannot find item with provided decimal number."})
                # блок выполнится, если не вызвалось исключение выше (айтем не был найден в БД)
                else:
                    # меняем децимальник у предка
                    assembly_to_change.decimal_number = assembly_decimal_number_to_edit
                    assembly_to_change.save()

                    # меняем поле первичная входимость у предков
                    for base_product in base_products_edit:
                        base_product_name = base_product['name']
                        base_product_entry_number = base_product['entry_number']
                        base_product_fields_to_edit = base_product['fields_to_edit']

                        base_product_db = BaseProduct.objects.filter(
                            name=base_product_name, entry_number=base_product_entry_number
                        )
                        for field in base_product_fields_to_edit:
                            base_product_db.field = base_product_fields_to_edit[field]
                        base_product_db.save()
            else:
                # случай если у сборки изменяется любое поле, кроме децимального номера
                for assemb in assemblies_edit:
                    assemb_decimal_number = assemb['decimal_number']
                    assemb_fields_to_edit = assemb['fields_to_edit']

                    assembly_db = Assembly.objects.filter(
                        decimal_number=assemb_decimal_number
                    )
                    for field in assemb_fields_to_edit:
                        assembly_db.field = assemb_fields_to_edit[field]
                    assembly_db.save()

        if not assemblies_edit:
            # изменяем данные у текущей сборки:
            if item_type == TypeEnum.ASSEMBLY:
                print(data['decimal_number'])
                assembly_item_to_change = Assembly.objects.filter(decimal_number=data['decimal_number']).get()
                fields_to_change = data['fields_to_edit']
                for field in fields_to_change:
                    setattr(assembly_item_to_change, field, fields_to_change[field])
                    assembly_item_to_change.save()
            else:
                print(data['decimal_number'])
                assembly_item_to_change = BaseProduct.objects.filter(decimal_number=data['decimal_number']).get()
                fields_to_change = data['fields_to_edit']
                for field in fields_to_change:
                    setattr(assembly_item_to_change, field, fields_to_change[field])
                    assembly_item_to_change.save()

            for base_product in base_products_edit:
                base_product_name = base_product['name']
                base_product_entry_number = base_product['entry_number']
                base_product_fields_to_edit = base_product['fields_to_edit']

                base_product_db = BaseProduct.objects.filter(
                    name=base_product_name, entry_number=base_product_entry_number
                ).get()
                for field in base_product_fields_to_edit:
                    setattr(base_product_db, field, base_product_fields_to_edit[field])
                    base_product_db.save()

    return JsonResponse({"status_code": 200, "message": "Item has been successfully edited."})


@login_required
@project_permissions_required(permissions=[ProjectPermissions.PERM_PROJECT_READ])
def export_page(request):
    current_user = UserProfile.objects.filter(user=request.user).get()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Products(created at {date} by {user}).xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
        user=current_user.firstname + ' ' + current_user.lastname,
    )
    resp = bd_to_dict_helper(check_empty_bd=False)
    tab = dict_to_table(resp)
    type_dict = {
        0: 'Сборка',
        1: 'Деталь',
        2: 'Стандартное изделие',
        3: 'Прочие изделия',
        'assembly': 'Сборка'}
    t = list(map(lambda x: type_dict[x[-1]], tab))
    c = 1
    for i in range(len(tab)):
        tab[i][-1] = t[i]
        tab[i].insert(0, c)
        c += 1
    cols = ['ID', 'Децимальный номер', 'Наименование', 'Первичная входимость', 'Принадлежность']
    tab.insert(0, cols)
    color_dict = {
        'Сборка': '93C47D',
        'Деталь': 'E2EFDA',
        'Стандартное изделие': 'FCE4D6',
        'Прочие изделия': 'D9E1F2'
    }
    wb = Workbook()
    ws = wb.active
    # Заполнение
    for row, r in zip(ws.iter_rows(min_row=1, max_row=len(tab), max_col=len(tab[0])), tab):
        for cell, v in zip(row, r):
            cell.value = v
    # Форматирование
    for cell in list(ws)[1]:
        cell.fill = PatternFill(start_color='38761D',
                                end_color='38761D',
                                fill_type='solid')
    for row in list(ws)[2:]:
        color = color_dict[row[-1].value]
        for cell in row:
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type='solid')
    wb.save(response)
    return response
